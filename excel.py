from copy import deepcopy
import calendar
from openpyxl import Workbook
from datetime import date, datetime, timedelta
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import Alignment
import holidays

def generate_monthly_work_report(*args):
        employee_name = input("Enter Employee Name: ")
        client_name = input("Enter Client: ")
        project_manager = input("Enter name of Project Manager: ")
        month_data = int(input("Enter Month in format (01-12): "))
        year_data = int(input("Enter Year: "))
        first, last = calendar.monthrange(year_data, month_data)
        print(first, last)
        leave_taken = []
        m = int(input("Enter number of Leave taken in the month: "))
        for i in range(0,m):
                dates_LT = input("Enter dates of Leaves taken in format (yy-mm-dd):")
                leave_taken.append(datetime.strptime(dates_LT,"%Y-%m-%d").date())
        holidays = []
        x = int(input("Enter number of holidays in the month: "))
        for i in range(0,x):
                dates_HD = input("Enter dates of Holidays in format (yy-mm-dd):")
                holidays.append(datetime.strptime(dates_HD,"%Y-%m-%d").date())
        extra_work = []
        n = int(input("Enter number of extra days work: "))
        for i in range(0,n):
                dates_EW = input("Enter dates of extra work in format (yy-mm-dd):")
                extra_work.append(datetime.strptime(dates_EW,"%Y-%m-%d").date())

        book = Workbook()
        sheet = book.active

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30

        weekly_holiday = ["Saturday", "Sunday"]

        d1 = date(year=year_data, month=month_data, day=1)
        d2 = date(year=year_data, month=month_data, day=last) + timedelta(days=1)
        delta = d2 - d1


        sheet.merge_cells('A1:E1')
        sheet['A1'].fill = PatternFill("solid", start_color="FFFF00")

        cell = sheet.cell(row=1, column=1)
        cell.value = 'Timesheet of month of June 2022'
        cell.alignment = Alignment(horizontal='center', vertical='center')

        today = date.today()
        employee = ['employee name', employee_name, 'Client name', client_name]
        sheet.append(employee)

        sheet.merge_cells('A3:E3')
        sheet['A3'].fill = PatternFill("solid", start_color="FFA07A")
        cell = sheet.cell(row=3, column=1)
        cell.value = 'CH- Client Holiday, LT- Leave Taken, WH- Weekly Holiday'
        cell.alignment = Alignment(horizontal='center', vertical='center')

        row_data = {
                'Date': None,
                'Day': None,
                'Total hours': None,
                'Particular': 'Python'
        }

        #pandas date range -> start date and end_date -> list []

        fieldnames = ['Date', 'Day', 'Total hours', 'Particular']
        sheet.append(fieldnames)
        sheet_data = []
        count_WH = 0
        count_WD = 0
        count_HD = 0
        count_LT = 0
        for i in range(delta.days):
                a = d1 + timedelta(days=i)
                b = a.strftime('%A')
                if b in weekly_holiday:
                        row_value = row_data.copy()
                        row_value["Date"] = a
                        row_value["Day"] = b
                        row_value["Total hours"] = 'WH'
                        count_WH = count_WH + 1
                elif a in holidays:
                        row_value = row_data.copy()
                        row_value["Date"] = a
                        row_value["Day"] = b
                        row_value["Total hours"] = "Holiday"
                        count_HD = count_HD + 1
                elif a in leave_taken:
                        row_value = row_data.copy()
                        row_value["Date"] = a
                        row_value["Day"] = b
                        row_value["Total hours"] = "LT"
                        count_LT = count_LT + 1
                else:
                        row_value = row_data.copy()
                        row_value["Date"] = a
                        row_value["Day"] = b
                        row_value["Total hours"] = 9
                        count_WD = count_WD + 1
                sheet_data.append(row_value)

        for each_data in sheet_data:
                sheet.append(list(each_data.values()))

        sheet.merge_cells('A36:E36')
        sheet['A36'].fill = PatternFill("solid", start_color="DE3163")
        cell = sheet.cell(row=36, column=1)
        cell.value = 'Approval'

        work1 = ["Total No.of Days Worked", count_WD, "Name of the Project Manager", project_manager]
        sheet.append(work1)

        work2 = ["Total Leaves Taken", count_LT, "Signature of the Candidate", employee_name]
        sheet.append(work2)

        work3 = ["No. of other Holidays", count_HD]
        sheet.append(work3)

        fill_row = []
        fill_column = 0
        for row in sheet.iter_rows():
                for cell in row:
                        if cell.value in weekly_holiday:
                                fill_row.append(cell.row)
        fill_column = cell.column
        for each_row in fill_row:
                for each_column in range(1,fill_column):
                        sheet.cell(row=each_row, column=each_column).fill = PatternFill("solid", start_color="424242")

        rows = range(1, 44)
        columns = range(1, 10)
        for row in rows:
                for col in columns:
                        sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        book.save("sample.xlsx")
generate_monthly_work_report()
