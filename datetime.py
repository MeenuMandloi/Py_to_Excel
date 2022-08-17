# from datetime import date, timedelta
#
# today = date.today()
# print("Today's date:", today)
# d1 = date(2022, 6, 1)
# d2 = date(2022, 6, 30)
# delta = d2 - d1
#
# for i in range(delta.days + 1):
#     print(d1 + timedelta(days=i))

from openpyxl import Workbook
from openpyxl.styles import PatternFill
# def background_colors("/home/my/Downloads/countries.csv"):
#     workbook = Workbook()
#     sheet = workbook.active
#     yellow = "00FFFF00"
#     for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=12):
#         for cell in rows:
#             if cell.row % 2:
#                 cell.fill = PatternFill(start_color=yellow, end_color=yellow,
#                                         fill_type = "solid")
#     workbook.save(path)
if __name__ == "__main__":
    background_colors("bg.xlsx")